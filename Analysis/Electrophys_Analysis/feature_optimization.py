# Optimize the features (for maximal performance using a linear regressor)

import os

import mne_bids
import numpy as np
import mne
import py_neuromodulation as nm
from py_neuromodulation import nm_analysis, nm_define_nmchannels, nm_plots, nm_settings, nm_decode
from bayes_opt import BayesianOptimization
from sklearn import metrics, model_selection, linear_model
from scipy.optimize import minimize
from scipy.stats import zscore
import matplotlib
from openpyxl import Workbook
from openpyxl import load_workbook
matplotlib.use('Qt5Agg')
import warnings
warnings.filterwarnings("ignore")

# Load the data
sub = "EL012"
path = f"..\\..\\..\\Data\\Off\\Neurophys\\Artifact_removal\\{sub}_cleaned.fif"
raw = mne.io.read_raw_fif(path).load_data()

sfreq = raw.info["sfreq"]
target_chan_name = raw.info["ch_names"][-1]
events = mne.events_from_annotations(raw)[0]

# Set channels
ch_names = ["ECOG_R_01_SMC_AT", "ECOG_R_02_SMC_AT", "ECOG_R_03_SMC_AT", "SPEED_MEAN"]
ch_types = ["ecog", "ecog", "ecog", "BEH"]
nm_channels = nm_define_nmchannels.set_channels(ch_names=ch_names, ch_types=ch_types, target_keywords="SPEED_MEAN")

# Attach the blocks together
blocks = []
for i in range(4):
    tmin = events[np.where((events[:, 2] == 2) | (events[:, 2] == 10002))[0], 0][int(96 * i)] / sfreq
    tmax = events[np.where((events[:, 2] == 1) | (events[:, 2] == 10001))[0], 0][int(96 * (i+1))-1] / sfreq
    block = raw.copy().crop(tmin=tmin, tmax=tmax).get_data(picks=ch_names)
    blocks.append(block)

blocks_all = np.hstack((blocks[0], blocks[1], blocks[2], blocks[3]))

# Save results in excel sheet
filename = "feature_optimization.xlsx"


def objective_function(samp_freq, s_seg, n_stack):

    # Set analysis parameters
    samp_freq = int(samp_freq)
    seg_ms = int(s_seg)

    # Settings
    settings = nm_settings.get_default_settings()
    settings = nm_settings.reset_settings(settings)
    settings["features"]["fft"] = True
    settings["features"]["return_raw"] = True
    settings["sampling_rate_features_hz"] = samp_freq
    settings["segment_length_features_ms"] = seg_ms
    settings["fft_settings"]["windowlength_ms"] = seg_ms
    del settings["frequency_ranges_hz"]["theta"]
    settings["postprocessing"]["feature_normalization"] = True
    settings["feature_normalization_settings"]["normalization_time_s"] = 1
    settings["feature_normalization_settings"]["normalization_method"] = "zscore"

    # Compute features
    stream = nm.Stream(
                settings=settings,
                nm_channels=nm_channels,
                verbose=False,
                sfreq=sfreq,
                line_noise=50
            )

    # Run the optimization
    perfs_all = []
    for i in range(5):

        if i < 4:
            data = blocks[i]
        else:
            data = blocks_all

        features = stream.run(data=data, out_path_root="..\\..\\..\\Data\\Off\\processed_data\\", folder_name=f"feature_optimization")
        feature_reader = nm_analysis.FeatureReader(
            feature_dir="..\\..\\..\\Data\\Off\\processed_data\\",
            feature_file="feature_optimization",
        )

        # Set the label
        feature_reader.label_name = "SPEED_MEAN"
        feature_reader.label = feature_reader.feature_arr[feature_reader.label_name]

        # Normalize target variable
        #feature_reader.label = zscore(feature_reader.label)

        # Setup the model and train
        model = linear_model.LinearRegression()

        try:
            feature_reader.decoder = nm_decode.Decoder(
                features=feature_reader.feature_arr,
                label=feature_reader.label,
                label_name=feature_reader.label_name,
                used_chs=feature_reader.used_chs,
                STACK_FEATURES_N_SAMPLES=True,
                time_stack_n_samples=int(n_stack),
                model=model,
                eval_method=metrics.r2_score,
                cv_method=model_selection.KFold(n_splits=4),
                VERBOSE=True
            )
            performances = feature_reader.run_ML_model(
                estimate_channels=True,
                estimate_gridpoints=False,
                estimate_all_channels_combined=True,
                save_results=True,
            )
            df_per = feature_reader.get_dataframe_performances(performances)
            perf = np.array(df_per["performance_test"])[-1]
        except Exception as e:
            perf = np.nan
            print(e)

        perfs_all.append(perf)

    # Calculate the average performance
    mean_perf = np.nanmean(perfs_all)
    perfs_all.append(mean_perf)
    row = [samp_freq, s_seg, n_stack] + perfs_all

    # Save decoding performance in excel sheet
    # Confirm file exists.
    # If not, create it, add headers, then append new data
    try:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]  # select first worksheet
    except FileNotFoundError:
        headers_row = ['samp_freq', 's_seg', 'n_stack', 'Block 1', 'Block 2', 'Block 3', 'Block 4', 'all', 'mean']
        wb = Workbook()
        ws = wb.active
        ws.append(headers_row)

    ws.append(row)
    wb.save(filename)

    return mean_perf


# Run the optimization
optimizer = BayesianOptimization(
    f=objective_function,
    pbounds={'samp_freq': (10, 60), 's_seg': (100, 500), 'n_stack': (5, 20)},
    verbose=2,
    random_state=1,
)

optimizer.maximize(
    init_points=10,
    n_iter=100,
)

for i, res in enumerate(optimizer.res):
    print("Iteration {}: \n\t{}".format(i, res))