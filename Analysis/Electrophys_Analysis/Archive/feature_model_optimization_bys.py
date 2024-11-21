# Optimize the features using a bayesian hyperparameter search and nested-cross validation
# Speed up the calculation by parallelizing the computation for all channel combinations
# Run this on ICN 1

import os
from joblib import Parallel, delayed
import mne_bids
import numpy as np
import mne
import py_neuromodulation as nm
from bayes_opt import BayesianOptimization
from sklearn import metrics, model_selection, linear_model
from sklearn.model_selection import KFold
import pandas as pd
from scipy.stats import zscore
from catboost import CatBoostRegressor
import matplotlib
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import pickle
import random
#sys.path.insert(1, "C:/CODE/ac_toolbox/")
import utils as u
#matplotlib.use('Qt5Agg')
import warnings
warnings.filterwarnings("ignore")
random.seed(420)


# Define parameters
n_outer = 6
n_inner = 4
n_optimization_rounds = 20
n_iterations_cb = 30

# Load the data
path = f"EL012.fif"
raw = mne.io.read_raw_fif(path).load_data()

sfreq = raw.info["sfreq"]
target_chan_name = raw.info["ch_names"][-1]
events = mne.events_from_annotations(raw)[0]

# Add ECoG channels with common average reference
ecog_names = ["ECOG_R_1_CAR", "ECOG_R_2_CAR", "ECOG_R_3_CAR"]
og_chan_names = ["ECOG_R_01_SMC_AT", "ECOG_R_02_SMC_AT", "ECOG_R_03_SMC_AT"]
for i, chan in enumerate(og_chan_names):
    new_ch = raw.get_data(chan) - raw.get_data(og_chan_names).mean(axis=0)
    u.add_new_channel(raw, new_ch, ecog_names[i], type="ecog")

# Add the LFP channel
lfp_names = ["LFP_1", "LFP_2"]
og_chan_names = ["LFP_R_01_STN_MT", "LFP_R_08_STN_MT"]
for i, chan in enumerate(og_chan_names):
    new_ch = raw.get_data(["LFP_R_02_STN_MT", "LFP_R_03_STN_MT", "LFP_R_04_STN_MT"]).sum(axis=0) - raw.get_data(chan)
    u.add_new_channel(raw, new_ch, lfp_names[i], type="ecog")

# Add a channel which marks the peak speed
peaks_idx = events[np.where((events[:, 2] == 3)), 0].flatten()
peaks_idx_ext = np.array([np.arange(x-60, x+60) for x in peaks_idx]).flatten()
peaks = np.zeros(raw._data.shape[-1])
peaks[peaks_idx_ext] = 1
u.add_new_channel(raw, peaks[np.newaxis, :], "PEAKS", type="misc")

# Define all channels and channel combinations to test
settings_decoding = ["ECoG_combined", "ECoG_LFP_combined", "LFP_combined"] + ecog_names + lfp_names


def run_nested_cross_val(setting):

    # Save results in excel sheet
    filename = f"decoding_results/feature_model_optimization_{setting}.xlsx"

    # Set channels
    if "combined" not in setting:
        ch_names = [setting, "SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog", "BEH", "BEH"]

    elif setting == "LFP_combined":
        ch_names = lfp_names + ["SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog"] * len(lfp_names) + ["BEH", "BEH"]

    elif setting == "ECoG_LFP_combined":
        ch_names = ecog_names + lfp_names + ["SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog"] * len(ecog_names + lfp_names) + ["BEH", "BEH"]

    elif setting == "ECoG_combined":
        ch_names = ecog_names + ["SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog"] * len(ecog_names) + ["BEH", "BEH"]

    channels = nm.utils.set_channels(
        ch_names=ch_names,
        ch_types=ch_types,
        reference=None,
        used_types=ch_types[:-2],
        target_keywords=["SPEED_MEAN", "PEAKS"],
    )

    # Attach the blocks together
    blocks = []
    for i in range(4):
        tmin = events[np.where((events[:, 2] == 2) | (events[:, 2] == 10002))[0], 0][int(96 * i)] / sfreq
        tmax = events[np.where((events[:, 2] == 1) | (events[:, 2] == 10001))[0], 0][int(96 * (i + 1)) - 1] / sfreq
        block = raw.copy().crop(tmin=tmin, tmax=tmax).get_data(picks=ch_names)
        blocks.append(block)
    blocks_all = np.hstack((blocks[0], blocks[1], blocks[2], blocks[3]))

    # Split the data into test and train set
    kf = KFold(n_splits=n_outer)
    train_idx_all = []; test_idx_all = []
    for train_index, test_index in kf.split(blocks_all.T):
        train_idx_all.append(train_index)
        test_idx_all.append(test_index)

    # Loop over the outer folds
    for count, n in enumerate(range(3, n_outer)):

        def objective_function(samp_freq, seg_ms, n_stack, learning_rate, depth):
            # Set analysis parameters
            samp_freq = int(samp_freq)
            seg_ms = int(seg_ms)

            # Compute performance on the test set using the optimal parameters
            settings = nm.NMSettings.get_fast_compute()
            settings.features.fft = True
            settings.features.return_raw = True
            settings.features.raw_hjorth = False
            settings.sampling_rate_features_hz = samp_freq
            settings.segment_length_features_ms = seg_ms
            settings.fft_settings.windowlength_ms = seg_ms
            del settings.frequency_ranges_hz["theta"]
            settings.postprocessing.feature_normalization = True
            settings.feature_normalization_settings.normalization_time_s = 1
            settings.feature_normalization_settings.normalization_method = "zscore"
            settings.preprocessing = settings.preprocessing[:2]

            # Compute features
            stream = nm.Stream(
                settings=settings,
                channels=channels,
                verbose=False,
                sfreq=sfreq,
                line_noise=50
            )

            data = blocks_all[:, train_idx_all[n]]

            stream.run(data=data, out_dir=f"decoding_results/", experiment_name=f"optimization_{setting}")
            feature_reader = nm.analysis.FeatureReader(
                feature_dir=f"decoding_results\\",
                feature_file=f"optimization_{setting}"
            )

            # Set the label
            feature_reader.label_name = "SPEED_MEAN"
            feature_reader.label = feature_reader.feature_arr[feature_reader.label_name]

            # Setup the model and train
            model = CatBoostRegressor(iterations=n_iterations_cb,
                                      depth=int(depth),
                                      learning_rate=learning_rate
                                      )

            try:
                feature_reader.decoder = nm.analysis.Decoder(
                    features=feature_reader.feature_arr,
                    label=feature_reader.label,
                    label_name=feature_reader.label_name,
                    used_chs=feature_reader.used_chs,
                    STACK_FEATURES_N_SAMPLES=True,
                    time_stack_n_samples=int(n_stack),
                    model=model,
                    eval_method=metrics.r2_score,
                    cv_method=model_selection.KFold(n_splits=n_inner),
                    VERBOSE=True
                )
                performances = feature_reader.run_ML_model(
                    estimate_channels=True,
                    estimate_gridpoints=False,
                    estimate_all_channels_combined=True,
                    save_results=True,
                )
                df_per = feature_reader.get_dataframe_performances(performances)
                perf = np.array(df_per["performance_test"])[-1]

                # Save decoding model
                #best_decoder_path = f"decoding_results/model_{setting}_fold_{n}_{samp_freq}_{seg_ms}_{n_stack}_{learning_rate}_{depth}.p"

                #with open(best_decoder_path, "wb") as output:
                #    pickle.dump(feature_reader.decoder, output)

            except Exception as e:
                    perf = 0
                    print(e)

            # Save
            row = [int(samp_freq), int(seg_ms), int(n_stack), int(depth), learning_rate] + [perf]

            # Save decoding performance in excel sheet
            headers_row = ['samp_freq', 'seg_ms', 'n_stack', 'depth', 'learning_rate', 'all']
            try:
                wb = load_workbook(filename)
                try:
                    ws = wb.worksheets[count]
                except:
                    wb.create_sheet(f"Fold {n}")
                    ws = wb.worksheets[count]
                    ws.append(headers_row)
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
                ws.title = f"Fold {n}"
                ws.append(headers_row)

            ws.append(row)
            wb.save(filename)

            return perf

        # Run the optimization
        optimizer = BayesianOptimization(
            f=objective_function,
            pbounds={'samp_freq': (20, 50), 'seg_ms': (200, 500), 'n_stack': (10, 20), 'learning_rate': (0.001,  1), 'depth': (4, 10)},
            verbose=2,
            random_state=1,
        )

        optimizer.maximize(
            init_points=n_optimization_rounds,
            n_iter=3,
        )

        # Get the optimal parameters yielding the highest performance
        df = pd.read_excel(filename, sheet_name=f"Fold {count}")
        samp_freq, seg_ms, n_stack, depth, learning_rate, _ = df.loc[df['all'].idxmax()]

        # Compute performance on the test set using the optimal parameters
        settings = nm.NMSettings.get_fast_compute()
        settings.features.fft = True
        settings.features.return_raw= True
        settings.sampling_rate_features_hz = samp_freq
        settings.segment_length_features_ms = seg_ms
        settings.fft_settings.windowlength_ms = seg_ms
        del settings.frequency_ranges_hz.theta
        settings.postprocessing.feature_normalization = True
        settings.feature_normalization_settings.normalization_time_s = 1
        settings.feature_normalization_settings.normalization_method = "zscore"

        # Compute features
        stream = nm.Stream(
            settings=settings,
            channels=channels,
            verbose=False,
            sfreq=sfreq,
            line_noise=50
        )

        data_train = blocks_all[:, train_idx_all[n]]
        data_test = blocks_all[:, test_idx_all[n]]
        features_train = stream.run(data=data_train)
        features_test = stream.run(data=data_test)

        # Get train and test set
        X_train = np.array(features_train)[:, :-3]
        X_test = np.array(features_test)[:, :-3]
        y_train = np.array(features_train.SPEED_MEAN)
        y_test = np.array(features_test.SPEED_MEAN)
        X_train_long, y_train = u.append_previous_n_samples(X=X_train,y=y_train, n=int(n_stack))
        X_test_long, y_test = u.append_previous_n_samples(X=X_test,y=y_test, n=int(n_stack))

        # Setup the model and train
        model = CatBoostRegressor(iterations=n_iterations_cb,
                                  depth=int(depth),
                                  learning_rate=learning_rate
                                  )

        # Fit the model
        model.fit(X_train_long, y_train)

        # Evaluate on test set
        y_test_pr = model.predict(X_test_long)
        perf = metrics.r2_score(y_test, y_test_pr)

        # Save performance with parameters in new excel worksheet
        row = [int(samp_freq), int(seg_ms), int(n_stack), int(depth), learning_rate] + [perf]

        # Save decoding performance in excel sheet
        wb = load_workbook(filename)
        ws = wb.worksheets[count]
        ws.append(row)
        wb.save(filename)


if __name__ == "__main__":
    #run_nested_cross_val(settings_decoding[0])
    #Parallel(n_jobs=len(settings_decoding))(delayed(run_nested_cross_val)(setting) for setting in settings_decoding)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Fold"
    ws.append([1, 2, 3])
    wb.save("test.xlsx")