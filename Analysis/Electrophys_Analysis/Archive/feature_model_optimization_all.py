# Optimize the features (for maximal performance using a linear regressor)

import os

import mne_bids
import numpy as np
import mne
import py_neuromodulation as nm
from py_neuromodulation import nm_analysis, nm_define_nmchannels, nm_plots, nm_settings, nm_decode
from bayes_opt import BayesianOptimization
from sklearn import metrics, model_selection, linear_model
import pandas as pd
from scipy.stats import zscore
from catboost import CatBoostRegressor
import matplotlib
from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import pickle
import random
sys.path.insert(1, "C:/CODE/ac_toolbox/")
import utils as u
matplotlib.use('Qt5Agg')
import warnings
warnings.filterwarnings("ignore")
random.seed(420)

# Load the data
sub = "EL012"
path = f"..\\..\\..\\Data\\Off\\Neurophys\\Artifact_removal\\{sub}_cleaned.fif"
raw = mne.io.read_raw_fif(path).load_data()

sfreq = raw.info["sfreq"]
target_chan_name = raw.info["ch_names"][-1]
events = mne.events_from_annotations(raw)[0]

# Add ECoG channels with common average reference
ecog_names = ["ECOG_R_1_CAR", "ECOG_R_2_CAR", "ECOG_R_3_CAR"]
og_chan_names = ["ECOG_R_01_SMC_AT", "ECOG_R_02_SMC_AT", "ECOG_R_03_SMC_AT"]
for i, chan in enumerate(og_chan_names):
    new_ch = raw.get_data(chan) - raw.get_data(og_chan_names).mean(axis=0)
    u.add_new_channel(raw, new_ch, ecog_names[i], type="ecog")

# Add the LFP channel
lfp_names = ["LFP_1", "LFP_2"]
og_chan_names = ["LFP_R_01_STN_MT", "LFP_R_08_STN_MT"]
for i, chan in enumerate(og_chan_names):
    new_ch = raw.get_data(["LFP_R_02_STN_MT", "LFP_R_03_STN_MT", "LFP_R_04_STN_MT"]).sum(axis=0) - raw.get_data(chan)
    u.add_new_channel(raw, new_ch, lfp_names[i], type="ecog")

# Add a channel which marks the peak speed
peaks_idx = events[np.where((events[:, 2] == 3)), 0].flatten()
peaks_idx_ext = np.array([np.arange(x-60, x+60) for x in peaks_idx]).flatten()
peaks = np.zeros(raw._data.shape[-1])
peaks[peaks_idx_ext] = 1
u.add_new_channel(raw, peaks[np.newaxis, :], "PEAKS", type="misc")

settings_decoding = ecog_names + lfp_names + ["ECoG_combined", "LFP_combined", "ECoG_LFP_combined"]

# Save results in excel sheet
filename = "feature_model_optimization2.xlsx"

for s, setting in enumerate(settings_decoding[3:]):

    # Set channels
    if "combined" not in setting:
        ch_names = [setting, "SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog", "BEH", "BEH"]

    elif setting == "LFP_combined":
        ch_names = lfp_names + ["SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog"] * len(lfp_names) + ["BEH", "BEH"]

    elif setting == "ECoG_LFP_combined":
        ch_names = ecog_names + lfp_names + ["SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog"] * len(ecog_names+lfp_names) + ["BEH", "BEH"]

    elif setting == "ECoG_combined":
        ch_names = ecog_names + ["SPEED_MEAN", "PEAKS"]
        ch_types = ["ecog"] * len(ecog_names) + ["BEH", "BEH"]

    nm_channels = nm_define_nmchannels.set_channels(ch_names=ch_names, ch_types=ch_types, target_keywords=["SPEED_MEAN", "PEAKS"], reference=None)

    # Attach the blocks together
    blocks = []
    for i in range(4):
        tmin = events[np.where((events[:, 2] == 2) | (events[:, 2] == 10002))[0], 0][int(96 * i)] / sfreq
        tmax = events[np.where((events[:, 2] == 1) | (events[:, 2] == 10001))[0], 0][int(96 * (i+1))-1] / sfreq
        block = raw.copy().crop(tmin=tmin, tmax=tmax).get_data(picks=ch_names)
        blocks.append(block)

    blocks_all = np.hstack((blocks[0], blocks[1], blocks[2], blocks[3]))

    def objective_function(samp_freq, s_seg, n_stack, learning_rate, depth):

        # Set analysis parameters
        samp_freq = int(samp_freq)
        seg_ms = int(s_seg)

        # Settings
        settings = nm_settings.get_default_settings()
        settings = nm_settings.reset_settings(settings)
        settings["features"]["fft"] = True
        settings["features"]["return_raw"] = True
        settings["sampling_rate_features_hz"] = samp_freq
        settings["segment_length_features_ms"] = seg_ms
        settings["fft_settings"]["windowlength_ms"] = seg_ms
        del settings["frequency_ranges_hz"]["theta"]
        settings["postprocessing"]["feature_normalization"] = True
        settings["feature_normalization_settings"]["normalization_time_s"] = 1
        settings["feature_normalization_settings"]["normalization_method"] = "zscore"

        # Compute features
        stream = nm.Stream(
                    settings=settings,
                    nm_channels=nm_channels,
                    verbose=False,
                    sfreq=sfreq,
                    line_noise=50
                )

        # Run the optimization
        data = blocks_all

        features = stream.run(data=data, out_path_root="..\\..\\..\\Data\\Off\\processed_data\\", folder_name=f"feature_optimization")
        feature_reader = nm_analysis.FeatureReader(
            feature_dir="..\\..\\..\\Data\\Off\\processed_data\\",
            feature_file="feature_optimization",
        )

        # Set the label
        feature_reader.label_name = "SPEED_MEAN"
        feature_reader.label = feature_reader.feature_arr[feature_reader.label_name]

        # Normalize target variable
        #feature_reader.label = zscore(feature_reader.label)

        # Setup the model and train
        model = CatBoostRegressor(iterations=50,
                                  depth=int(depth),
                                  learning_rate=learning_rate
                                  )

        try:
            feature_reader.decoder = nm_decode.Decoder(
                features=feature_reader.feature_arr,
                label=feature_reader.label,
                label_name=feature_reader.label_name,
                used_chs=feature_reader.used_chs,
                STACK_FEATURES_N_SAMPLES=True,
                time_stack_n_samples=int(n_stack),
                model=model,
                eval_method=metrics.r2_score,
                cv_method=model_selection.KFold(n_splits=4),
                VERBOSE=True
            )
            performances = feature_reader.run_ML_model(
                estimate_channels=True,
                estimate_gridpoints=False,
                estimate_all_channels_combined=True,
                save_results=True,
            )
            df_per = feature_reader.get_dataframe_performances(performances)
            perf = np.array(df_per["performance_test"])[-1]

        except Exception as e:
            perf = 0
            print(e)

        # Save
        row = [samp_freq, s_seg, n_stack, depth, learning_rate] + [perf]

        # Save decoding performance in excel sheet
        # Confirm file exists.
        # If not, create it, add headers, then append new data
        headers_row = ['samp_freq', 's_seg', 'n_stack', 'depth', 'learning_rate', 'all']
        try:
            wb = load_workbook(filename)
            try:
                ws = wb.worksheets[s]  # select first worksheet
                df = pd.read_excel(filename, sheet_name=setting)
                max_before = df["all"].max()
            except:
                wb.create_sheet(setting)
                ws = wb.worksheets[s]
                ws.append(headers_row)
                max_before = 0
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = setting
            ws.append(headers_row)
            max_before = 0

        ws.append(row)
        wb.save(filename)

        # If highest performance, save with special filename
        if max_before < perf:
            best_decoder_path = f"best_model_{setting}.p"
            with open(best_decoder_path, "wb") as output:  # Overwrites any existing file.
                pickle.dump(feature_reader.decoder, output)

        return perf


    # Run the optimization
    optimizer = BayesianOptimization(
        f=objective_function,
        pbounds={'samp_freq': (20, 50), 's_seg': (200, 500), 'n_stack': (10, 20), 'learning_rate': (0.001,  1), 'depth': (4, 10)},
        verbose=2,
        random_state=1,
    )

    optimizer.maximize(
        init_points=20,
        n_iter=5,
    )

    for i, res in enumerate(optimizer.res):
        print("Iteration {}: \n\t{}".format(i, res))