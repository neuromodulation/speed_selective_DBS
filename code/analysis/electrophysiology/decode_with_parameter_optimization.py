# Decode contineous movement speed using catboost machine learning model 
# Optimize feature extraction and model parameters uisng bayesian optimization
# Run on BIH cluster with parallel processing

# Prepare environment
import numpy as np
import mne
import py_neuromodulation as nm
from bayes_opt import BayesianOptimization
from sklearn import metrics, model_selection
from sklearn.model_selection import KFold
import pandas as pd
from catboost import CatBoostRegressor
from openpyxl import Workbook
from openpyxl import load_workbook
import random
from joblib import Parallel, delayed
from itertools import product
import shutil
random.seed(420)

# Define helper functions
def append_previous_n_samples(X: np.ndarray, y: np.ndarray, n: int = 5):
    TIME_DIM = X.shape[0] - n
    FEATURE_DIM = int(n * X.shape[1])
    time_arr = np.empty((TIME_DIM, FEATURE_DIM))
    for time_idx, time_ in enumerate(np.arange(n, X.shape[0])):
        for time_point in range(n):
            time_arr[
                time_idx,
                time_point * X.shape[1] : (time_point + 1) * X.shape[1],
            ] = X[time_ - time_point, :]
    return time_arr, y[n:]

# Run validation for one outer cross-validation fold
def run_nested_cross_val(setting, n):

    # Save results in excel sheet
    filename = f"../data/electrophysiology/decoding_results/feature_model_optimization_{setting}_{n}.xlsx"

    # Set channels
    if "combined" not in setting:
        ch_names = [setting, "SPEED_MEAN"]
        ch_types = ["ecog", "BEH"]

    elif setting == "LFP_combined":
        ch_names = lfp_names + ["SPEED_MEAN"]
        ch_types = ["ecog"] * len(lfp_names) + ["BEH"]

    elif setting == "ECoG_LFP_combined":
        ch_names = ecog_names + lfp_names + ["SPEED_MEAN"]
        ch_types = ["ecog"] * len(ecog_names + lfp_names) + ["BEH"]

    elif setting == "ECoG_combined":
        ch_names = ecog_names + ["SPEED_MEAN"]
        ch_types = ["ecog"] * len(ecog_names) + ["BEH"]

    channels = nm.utils.set_channels(
        ch_names=ch_names,
        ch_types=ch_types,
        reference=None,
        used_types=ch_types[:-1],
        target_keywords=["SPEED_MEAN"],
    )

    # Attach the blocks together (electrophysiology was continuously recorded, but speed only during the blocks and not during the breaks)
    blocks = []
    for i in range(4):
        tmin = events[np.where((events[:, 2] == 2) | (events[:, 2] == 10002))[0], 0][int(96 * i)] / sfreq
        tmax = events[np.where((events[:, 2] == 1) | (events[:, 2] == 10001))[0], 0][int(96 * (i + 1)) - 1] / sfreq
        block = raw.copy().crop(tmin=tmin, tmax=tmax).get_data(picks=ch_names)
        blocks.append(block)
    blocks_all = np.hstack((blocks[0], blocks[1], blocks[2], blocks[3]))

    # Split the data into test and train set
    kf = KFold(n_splits=n_outer)
    train_idx_all = []; test_idx_all = []
    for train_index, test_index in kf.split(blocks_all.T):
        train_idx_all.append(train_index)
        test_idx_all.append(test_index)

    # Function whose parameters are optimized in the hyperparameter search
    # Feature sampling frequency, feature window, number of stacked samples, catboost learning rate and tree depth)
    def objective_function(samp_freq, seg_ms, n_stack, learning_rate, depth):

        # Set analysis parameters
        samp_freq = int(samp_freq)
        seg_ms = int(seg_ms)

        # Compute the accuracy on the test set using the optimal parameters
        settings = nm.NMSettings.get_fast_compute()
        settings.features.fft = True
        settings.features.return_raw = True
        settings.features.raw_hjorth = False
        settings.sampling_rate_features_hz = samp_freq
        settings.segment_length_features_ms = seg_ms
        settings.fft_settings.windowlength_ms = seg_ms
        del settings.frequency_ranges_hz["theta"]
        settings.postprocessing.feature_normalization = True
        settings.feature_normalization_settings.normalization_time_s = 1
        settings.feature_normalization_settings.normalization_method = "zscore"
        settings.preprocessing = settings.preprocessing[:2]

        # Compute features
        stream = nm.Stream(
            settings=settings,
            channels=channels,
            verbose=False,
            sfreq=sfreq,
            line_noise=50
        )

        data = blocks_all[:, train_idx_all[n]]

        stream.run(data=data, out_dir="../data/electrophysiology/features/", experiment_name=f"optimization_{setting}_{n}")
        feature_reader = nm.analysis.FeatureReader(
            feature_dir="../data/electrophysiology/features/",
            feature_file=f"optimization_{setting}_{n}"
        )

        # Set the label
        feature_reader.label_name = "SPEED_MEAN"
        feature_reader.label = feature_reader.feature_arr[feature_reader.label_name]

        # Setup the model and train
        model = CatBoostRegressor(iterations=n_iterations_cb,
                                  depth=int(depth),
                                  learning_rate=learning_rate
                                  )

        feature_reader.decoder = nm.analysis.Decoder(
            features=feature_reader.feature_arr,
            label=feature_reader.label,
            label_name=feature_reader.label_name,
            used_chs=feature_reader.used_chs,
            STACK_FEATURES_N_SAMPLES=True,
            time_stack_n_samples=int(n_stack),
            model=model,
            eval_method=metrics.r2_score,
            cv_method=model_selection.KFold(n_splits=n_inner),
            VERBOSE=True
        )
        
        # Extract test accuracy
        estimate_combined = True if "combined" in setting else False
        estimate_channels = False if "combined" in setting else True
        performances = feature_reader.run_ML_model(
            estimate_channels=estimate_channels,
            estimate_gridpoints=False,
            estimate_all_channels_combined=estimate_combined,
            save_results=True,
        )
        df_per = feature_reader.get_dataframe_performances(performances)
        acc = np.array(df_per["performance_test"])[-1]

        # Save
        row = [int(samp_freq), int(seg_ms), int(n_stack), int(depth), learning_rate] + [acc]

        # Save decoding accuracy in excel sheet
        headers_row = ['samp_freq', 'seg_ms', 'n_stack', 'depth', 'learning_rate', 'all']
        try:
            wb = load_workbook(filename)
            ws = wb.worksheets[0]
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Fold {n}"
            ws.append(headers_row)

        ws.append(row)
        wb.save(filename)
        
        # Delete the feature folder to avoid memory problems
        shutil.rmtree(f"../data/electrophysiology/features/optimization_{setting}_{n}/")

        return perf

    # Run the optimization
    optimizer = BayesianOptimization(
        f=objective_function,
        pbounds={'samp_freq': (20, 50), 'seg_ms': (200, 500), 'n_stack': (10, 20), 'learning_rate': (0.001,  1), 'depth': (4, 10)},
        verbose=2,
        random_state=1,
    )

    optimizer.maximize(
        init_points=n_optimization_rounds,
        n_iter=3,
    )

    # Get the optimal parameters yielding the highest accuracy
    df = pd.read_excel(filename, sheet_name=f"Fold {n}")
    samp_freq, seg_ms, n_stack, depth, learning_rate, _ = df.loc[df['all'].idxmax()]

    # Compute the accuracy on the test set using the optimal parameters
    settings = nm.NMSettings.get_fast_compute()
    settings.features.fft = True
    settings.features.return_raw = True
    settings.features.raw_hjorth = False
    settings.sampling_rate_features_hz = samp_freq
    settings.segment_length_features_ms = seg_ms
    settings.fft_settings.windowlength_ms = seg_ms
    del settings.frequency_ranges_hz["theta"]
    settings.postprocessing.feature_normalization = True
    settings.feature_normalization_settings.normalization_time_s = 1
    settings.feature_normalization_settings.normalization_method = "zscore"
    settings.preprocessing = settings.preprocessing[:2]

    # Compute features
    stream = nm.Stream(
        settings=settings,
        channels=channels,
        verbose=False,
        sfreq=sfreq,
        line_noise=50
    )

    data_train = blocks_all[:, train_idx_all[n]]
    data_test = blocks_all[:, test_idx_all[n]]
    features_train = stream.run(data=data_train)
    features_test = stream.run(data=data_test)

    # Get train and test set (without time and label)
    X_train = np.array(features_train)[:, :-2]
    X_test = np.array(features_test)[:, :-2]
    y_train = np.array(features_train.SPEED_MEAN)
    y_test = np.array(features_test.SPEED_MEAN)
    X_train_long, y_train = append_previous_n_samples(X=X_train,y=y_train, n=int(n_stack))
    X_test_long, y_test = append_previous_n_samples(X=X_test,y=y_test, n=int(n_stack))

    # Setup the model and train
    model = CatBoostRegressor(iterations=n_iterations_cb,
                              depth=int(depth),
                              learning_rate=learning_rate
                              )

    # Fit the model
    model.fit(X_train_long, y_train)

    # Evaluate on test set
    y_test_pr = model.predict(X_test_long)
    acc = metrics.r2_score(y_test, y_test_pr)

    # Save accuracy with parameters in new excel worksheet
    row = [int(samp_freq), int(seg_ms), int(n_stack), int(depth), learning_rate] + [acc]

    # Save decoding accuracy in excel sheet
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    ws.append(row)
    wb.save(filename)


if __name__ == "__main__":
    # Define parameters
    n_outer = 8
    n_inner = 4
    n_optimization_rounds = 30
    n_iterations_cb = 30

    # Load the data
    path = "../data/electrophysiology/ECoG_LFP_for_decoding.fif"
    raw = mne.io.read_raw_fif(path).load_data()
    sfreq = raw.info["sfreq"]
    ch_names = raw.info["ch_names"]
    ecog_names = [name for name in ch_names if "ECOG" in name]
    lfp_names = [name for name in ch_names if "LFP" in name]
    events = mne.events_from_annotations(raw)[0]

    # Define all channels and channel combinations to test
    settings_decoding = ["ECoG_combined", "ECoG_LFP_combined", "LFP_combined"] + ecog_names + lfp_names

    # Run parallelized on cluster (one setting and outer fold separate)
    l1, l2 = range(len(settings_decoding)), range(n_outer)
    runs = list(product(l1, l2))

    # On BIH cluster
    #input1 = int(sys.argv[1]) - 1
    #run_nested_cross_val(settings_decoding[runs[input1][0]], runs[input1][1])

    # On laptop
    Parallel(n_jobs=1)(delayed(run_nested_cross_val)(settings_decoding[setting_idx], outer_fold_idx) for setting_idx, outer_fold_idx in runs)